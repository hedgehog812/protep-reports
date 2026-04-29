import os
from datetime import datetime
from functools import wraps
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, flash, redirect, render_template, request, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-key-change-later")

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "reports")

TEMP_DIR = Path("generated_reports")
TEMP_DIR.mkdir(exist_ok=True)

USERS = {
    "admin": {
        "password": "admin123",
        "role": "admin",
        "display_name": "Администратор",
    },
    "manager": {
        "password": "manager123",
        "role": "manager",
        "display_name": "Менеджер заявок",
    },
    "user": {
        "password": "user123",
        "role": "user",
        "display_name": "Исполнитель",
    },
}

REQUEST_STATUSES = ["К выполнению", "В процессе", "Выполнено"]
REQUEST_PRIORITIES = ["Низкий", "Средний", "Высокий", "Критический"]


def db():
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Не заполнены SUPABASE_URL и SUPABASE_KEY.")
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


def role_required(allowed_roles):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if "username" not in session:
                return redirect(url_for("login"))
            if session.get("role") not in allowed_roles:
                flash("Недостаточно прав для выполнения операции.", "error")
                return redirect(url_for("index"))
            return func(*args, **kwargs)
        return wrapper
    return decorator


admin_required = role_required(["admin"])
manager_required = role_required(["admin", "manager"])
request_status_required = role_required(["admin", "manager", "user"])
report_generate_required = role_required(["admin", "manager"])
report_manage_required = role_required(["admin"])


def current_user():
    return {
        "username": session.get("username"),
        "role": session.get("role"),
        "display_name": session.get("display_name"),
    }


def fetch_table(name):
    return db().table(name).select("*").order("created_at", desc=True).execute().data


def fetch_contracts():
    return (
        db()
        .table("contracts")
        .select("*, objects(name, address)")
        .order("created_at", desc=True)
        .execute()
        .data
    )


def fetch_requests():
    return (
        db()
        .table("service_requests")
        .select("*, objects(name, address)")
        .order("created_at", desc=True)
        .execute()
        .data
    )


def fetch_completed_requests():
    return (
        db()
        .table("service_requests")
        .select("*, objects(name, address)")
        .eq("status", "Выполнено")
        .order("updated_at", desc=True)
        .execute()
        .data
    )


def group_requests_by_status(requests_data):
    grouped = {status: [] for status in REQUEST_STATUSES}
    for item in requests_data:
        grouped.setdefault(item.get("status", "К выполнению"), []).append(item)
    return grouped


def user_can_manage_records():
    return session.get("role") in ["admin", "manager"]


def user_can_manage_reports():
    return session.get("role") in ["admin"]


def user_can_generate_reports():
    return session.get("role") in ["admin", "manager"]


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        user = USERS.get(username)
        if not user or user["password"] != password:
            flash("Неверный логин или пароль.", "error")
            return redirect(url_for("login"))

        session["username"] = username
        session["role"] = user["role"]
        session["display_name"] = user["display_name"]

        return redirect(url_for("index"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    try:
        objects = fetch_table("objects")
        contracts = fetch_contracts()
        requests_data = fetch_requests()
        reports = fetch_table("reports")
    except Exception as error:
        objects, contracts, requests_data, reports = [], [], [], []
        flash(str(error), "error")

    return render_template(
        "index.html",
        objects=objects,
        contracts=contracts,
        service_requests=requests_data,
        requests_by_status=group_requests_by_status(requests_data),
        reports=reports,
        user=current_user(),
        request_statuses=REQUEST_STATUSES,
        request_priorities=REQUEST_PRIORITIES,
        can_manage_records=user_can_manage_records(),
        can_manage_reports=user_can_manage_reports(),
        can_generate_reports=user_can_generate_reports(),
    )


@app.route("/objects", methods=["POST"])
@manager_required
def add_object():
    try:
        db().table("objects").insert({
            "name": request.form["name"].strip(),
            "address": request.form["address"].strip(),
            "area": float(request.form["area"].replace(",", ".")),
            "status": request.form["status"],
            "responsible_person": request.form.get("responsible_person", "").strip(),
        }).execute()
        flash("Объект недвижимости добавлен.", "success")
    except Exception as error:
        flash(f"Ошибка при добавлении объекта: {error}", "error")
    return redirect(url_for("index"))


@app.route("/objects/<int:object_id>/delete", methods=["POST"])
@manager_required
def delete_object(object_id):
    try:
        db().table("objects").delete().eq("id", object_id).execute()
        flash("Объект недвижимости удалён. Связанные договоры и заявки также удалены.", "success")
    except Exception as error:
        flash(f"Ошибка при удалении объекта: {error}", "error")
    return redirect(url_for("index"))


@app.route("/contracts", methods=["POST"])
@manager_required
def add_contract():
    try:
        db().table("contracts").insert({
            "object_id": int(request.form["object_id"]),
            "tenant_name": request.form["tenant_name"].strip(),
            "contract_number": request.form["contract_number"].strip(),
            "monthly_payment": float(request.form["monthly_payment"].replace(",", ".")),
            "start_date": request.form["start_date"],
            "end_date": request.form["end_date"],
            "status": request.form["contract_status"],
        }).execute()
        flash("Договор добавлен.", "success")
    except Exception as error:
        flash(f"Ошибка при добавлении договора: {error}", "error")
    return redirect(url_for("index"))


@app.route("/contracts/<int:contract_id>/delete", methods=["POST"])
@manager_required
def delete_contract(contract_id):
    try:
        db().table("contracts").delete().eq("id", contract_id).execute()
        flash("Договор удалён.", "success")
    except Exception as error:
        flash(f"Ошибка при удалении договора: {error}", "error")
    return redirect(url_for("index"))


@app.route("/requests", methods=["POST"])
@manager_required
def add_request():
    try:
        db().table("service_requests").insert({
            "object_id": int(request.form["request_object_id"]),
            "title": request.form["title"].strip(),
            "description": request.form.get("description", "").strip(),
            "priority": request.form["priority"],
            "status": request.form.get("request_status", "К выполнению"),
            "created_by": session.get("username"),
            "updated_by": session.get("username"),
            "updated_at": datetime.now().isoformat(),
        }).execute()
        flash("Заявка добавлена.", "success")
    except Exception as error:
        flash(f"Ошибка при добавлении заявки: {error}", "error")
    return redirect(url_for("index"))


@app.route("/requests/<int:request_id>/update", methods=["POST"])
@manager_required
def update_request(request_id):
    try:
        title = request.form.get("title", "").strip()
        priority = request.form.get("priority", "").strip()
        status = request.form.get("status", "").strip()
        description = request.form.get("description", "").strip()

        if status not in REQUEST_STATUSES:
            flash("Некорректный статус заявки.", "error")
            return redirect(url_for("index"))

        if priority not in REQUEST_PRIORITIES:
            flash("Некорректный уровень срочности.", "error")
            return redirect(url_for("index"))

        db().table("service_requests").update({
            "title": title,
            "description": description,
            "priority": priority,
            "status": status,
            "updated_by": session.get("username"),
            "updated_at": datetime.now().isoformat(),
        }).eq("id", request_id).execute()

        flash("Заявка обновлена.", "success")
    except Exception as error:
        flash(f"Ошибка при обновлении заявки: {error}", "error")
    return redirect(url_for("index"))


@app.route("/requests/<int:request_id>/status", methods=["POST"])
@request_status_required
def change_request_status(request_id):
    try:
        new_status = request.form.get("status", "").strip()

        if new_status not in REQUEST_STATUSES:
            flash("Некорректный статус заявки.", "error")
            return redirect(url_for("index"))

        db().table("service_requests").update({
            "status": new_status,
            "updated_by": session.get("username"),
            "updated_at": datetime.now().isoformat(),
        }).eq("id", request_id).execute()

        flash("Статус заявки изменён.", "success")
    except Exception as error:
        flash(f"Ошибка при изменении статуса заявки: {error}", "error")
    return redirect(url_for("index"))


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="D9EAF7")
    side = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row >= 4:
                cell.border = border

    if ws.max_row >= 4:
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_excel_report(objects, contracts, service_requests):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_date = datetime.now().strftime("%d.%m.%Y")
    file_name = f"protep_daily_completed_requests_{timestamp}.xlsx"
    file_path = TEMP_DIR / file_name

    wb = Workbook()

    ws = wb.active
    ws.title = "Итоги дня"

    completed_requests = [r for r in service_requests if r.get("status") == "Выполнено"]
    critical_completed = [r for r in completed_requests if r.get("priority") == "Критический"]
    high_completed = [r for r in completed_requests if r.get("priority") == "Высокий"]

    ws["A1"] = "Дневной отчёт по выполненным заявкам АО «ПРОТЭП»"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:D1")

    rows = [
        ["Дата формирования", datetime.now().strftime("%d.%m.%Y %H:%M")],
        ["Отчётный день", report_date],
        ["Количество выполненных заявок", len(completed_requests)],
        ["Критических выполненных заявок", len(critical_completed)],
        ["Высокоприоритетных выполненных заявок", len(high_completed)],
        ["Количество объектов в системе", len(objects)],
        ["Количество договоров в системе", len(contracts)],
    ]

    for idx, row in enumerate(rows, 3):
        ws[f"A{idx}"] = row[0]
        ws[f"B{idx}"] = row[1]

    ws_done = wb.create_sheet("Выполненные заявки")
    ws_done.append([
        "№",
        "Объект",
        "Адрес объекта",
        "Тема заявки",
        "Описание",
        "Срочность",
        "Статус",
        "Создал",
        "Обновил",
        "Дата создания",
        "Дата обновления",
    ])
    ws_done.insert_rows(1, 3)
    ws_done["A1"] = "Заявки, выполненные за день"
    ws_done["A1"].font = Font(size=14, bold=True)
    ws_done.merge_cells("A1:K1")

    for i, r in enumerate(completed_requests, 1):
        obj = r.get("objects") or {}
        ws_done.append([
            i,
            obj.get("name", ""),
            obj.get("address", ""),
            r.get("title", ""),
            r.get("description", ""),
            r.get("priority", ""),
            r.get("status", ""),
            r.get("created_by", ""),
            r.get("updated_by", ""),
            (r.get("created_at") or "")[:10],
            (r.get("updated_at") or "")[:19].replace("T", " "),
        ])

    ws_objects = wb.create_sheet("Объекты")
    ws_objects.append(["№", "Название", "Адрес", "Площадь, м²", "Статус", "Ответственный"])
    ws_objects.insert_rows(1, 3)
    ws_objects["A1"] = "Объекты недвижимости"
    ws_objects["A1"].font = Font(size=14, bold=True)
    ws_objects.merge_cells("A1:F1")

    for i, o in enumerate(objects, 1):
        ws_objects.append([
            i,
            o.get("name", ""),
            o.get("address", ""),
            float(o.get("area") or 0),
            o.get("status", ""),
            o.get("responsible_person", ""),
        ])

    for sheet in [ws, ws_done, ws_objects]:
        style_sheet(sheet)
        for col_idx in range(1, sheet.max_column + 1):
            letter = get_column_letter(col_idx)
            sheet.column_dimensions[letter].width = 24

    wb.save(file_path)
    return file_name, file_path


def upload_report(file_name, file_path):
    client = db()
    storage_path = f"reports/{file_name}"

    with open(file_path, "rb") as f:
        client.storage.from_(SUPABASE_BUCKET).upload(
            path=storage_path,
            file=f.read(),
            file_options={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "upsert": "true",
            },
        )

    file_url = client.storage.from_(SUPABASE_BUCKET).get_public_url(storage_path)
    return file_url, storage_path


@app.route("/reports/generate", methods=["POST"])
@report_generate_required
def generate_report():
    try:
        objects = fetch_table("objects")
        contracts = fetch_contracts()
        completed_requests = fetch_completed_requests()

        if not completed_requests:
            flash("Нельзя сформировать дневной отчёт: в колонке «Выполнено» нет заявок.", "error")
            return redirect(url_for("index"))

        file_name, file_path = create_excel_report(objects, contracts, completed_requests)
        file_url, storage_path = upload_report(file_name, file_path)

        db().table("reports").insert({
            "file_name": file_name,
            "file_url": file_url,
            "storage_path": storage_path,
            "report_type": "Дневной отчёт по выполненным заявкам",
            "created_by": session.get("username"),
        }).execute()

        for item in completed_requests:
            db().table("service_requests").delete().eq("id", item["id"]).execute()

        flash("Дневной отчёт сформирован. Колонка «Выполнено» очищена.", "success")

    except Exception as error:
        flash(f"Ошибка при формировании отчёта: {error}", "error")

    return redirect(url_for("index"))


@app.route("/reports/<int:report_id>/update", methods=["POST"])
@report_manage_required
def update_report(report_id):
    try:
        report_type = request.form.get("report_type", "").strip() or "Дневной отчёт по выполненным заявкам"
        db().table("reports").update({
            "report_type": report_type,
        }).eq("id", report_id).execute()
        flash("Отчёт изменён.", "success")
    except Exception as error:
        flash(f"Ошибка при изменении отчёта: {error}", "error")
    return redirect(url_for("index"))


@app.route("/reports/<int:report_id>/delete", methods=["POST"])
@report_manage_required
def delete_report(report_id):
    try:
        client = db()
        report_result = client.table("reports").select("*").eq("id", report_id).single().execute()
        report = report_result.data

        if not report:
            flash("Отчёт не найден.", "error")
            return redirect(url_for("index"))

        storage_path = report.get("storage_path")

        if storage_path:
            try:
                client.storage.from_(SUPABASE_BUCKET).remove([storage_path])
            except Exception:
                pass

        client.table("reports").delete().eq("id", report_id).execute()
        flash("Отчёт удалён.", "success")
    except Exception as error:
        flash(f"Ошибка при удалении отчёта: {error}", "error")
    return redirect(url_for("index"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
